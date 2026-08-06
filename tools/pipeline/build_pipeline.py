#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
案件管理表(案件一覧 + ダッシュボード)の xlsx を生成する。
顧客・案件・ステータス・金額・開催日を一元管理し、ステータスはプルダウン、
ダッシュボードは数式で自動集計する。Excel / Google スプレッドシートで開ける。

    python3 build_pipeline.py            # -> 案件管理.xlsx
"""
import os, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))

STATUSES = ["問合せ", "ヒアリング", "見積提出", "受注", "開催準備",
            "開催済", "請求済", "入金済", "失注/保留"]
PLANS = ["機材レンタルのみ", "運営代行込み", "設営・撤収のみ", "相談中"]

HEADERS = ["案件ID", "受付日", "顧客名", "経路/業種", "案件内容", "主なコンテンツ",
           "プラン", "開催日", "会場", "ステータス", "見積(税込)", "請求(税込)",
           "入金", "担当", "次アクション", "備考"]

# 種まきデータ(新通エスピーは実在の継続案件。他はサンプル→備考に明記)
ROWS = [
    ["NV-2026-001", datetime.date(2025,11,1), "株式会社新通エスピー", "既存/イベント", "エアー遊具 月次レンタル", "エアー遊具",
     "機材レンタルのみ", datetime.date(2026,3,1), "各会場", "入金済", 330000, 330000, "済", "竹村", "4月度分の見積送付", "継続案件(R7.11〜)"],
    ["NV-2026-002", datetime.date(2026,7,20), "(サンプル)◯◯モール", "HP問合せ/商業施設", "夏休みキッズイベント", "キッズライセンスパーク＋エア遊具",
     "運営代行込み", datetime.date(2026,8,20), "セントラルコート", "見積提出", 495000, None, "", "竹村", "8/8までに返答予定", "★サンプル行(要差し替え)"],
    ["NV-2026-003", datetime.date(2026,7,28), "(サンプル)△△自動車 販売店", "紹介/店舗集客", "週末来店フェア", "スクイーズWS＋カーニバル",
     "機材レンタルのみ", datetime.date(2026,9,7), "店頭特設", "ヒアリング", None, None, "", "竹村", "会場図面を依頼", "★サンプル行(要差し替え)"],
]


def build():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "案件一覧"

    ink = "0B2545"; head_fill = PatternFill("solid", fgColor=ink)
    head_font = Font(name="Meiryo", bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="DFE7F2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, c, h)
        cell.fill = head_fill; cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for r, row in enumerate(ROWS, 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, v)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(c in (5, 6, 15, 16)))
            if c in (11, 12):  # 金額
                cell.number_format = '¥#,##0'
            if c in (2, 8):    # 日付
                cell.number_format = 'yyyy/m/d'

    # 幅
    widths = [12,11,20,14,20,22,15,11,14,11,12,12,7,9,20,22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    # 入力規則(プルダウン): ステータス(J), プラン(G), 入金(M)
    last = 400
    dv_status = DataValidation(type="list", formula1='"' + ",".join(STATUSES) + '"', allow_blank=True)
    dv_plan   = DataValidation(type="list", formula1='"' + ",".join(PLANS) + '"', allow_blank=True)
    dv_pay    = DataValidation(type="list", formula1='"未,一部,済"', allow_blank=True)
    ws.add_data_validation(dv_status); ws.add_data_validation(dv_plan); ws.add_data_validation(dv_pay)
    dv_status.add(f"J2:J{last}"); dv_plan.add(f"G2:G{last}"); dv_pay.add(f"M2:M{last}")

    # ダッシュボード
    dash = wb.create_sheet("ダッシュボード")
    dash["A1"] = "案件ダッシュボード"; dash["A1"].font = Font(name="Meiryo", bold=True, size=16, color=ink)
    dash["A3"] = "ステータス別 件数"; dash["A3"].font = Font(bold=True)
    for i, st in enumerate(STATUSES, 4):
        dash.cell(i, 1, st)
        dash.cell(i, 2, f'=COUNTIF(案件一覧!$J$2:$J${last},A{i})')
    dash["D3"] = "金額サマリー(税込)"; dash["D3"].font = Font(bold=True)
    summ = [
        ("パイプライン見込(見積合計)", '=SUM(案件一覧!$K$2:$K$'+str(last)+')'),
        ("受注以降の見積額", '=SUMIFS(案件一覧!$K$2:$K$'+str(last)+',案件一覧!$J$2:$J$'+str(last)+',"受注")+SUMIFS(案件一覧!$K$2:$K$'+str(last)+',案件一覧!$J$2:$J$'+str(last)+',"開催準備")+SUMIFS(案件一覧!$K$2:$K$'+str(last)+',案件一覧!$J$2:$J$'+str(last)+',"開催済")'),
        ("請求済 合計", '=SUM(案件一覧!$L$2:$L$'+str(last)+')'),
        ("入金済 合計", '=SUMIFS(案件一覧!$L$2:$L$'+str(last)+',案件一覧!$M$2:$M$'+str(last)+',"済")'),
        ("未入金 残", '=SUM(案件一覧!$L$2:$L$'+str(last)+')-SUMIFS(案件一覧!$L$2:$L$'+str(last)+',案件一覧!$M$2:$M$'+str(last)+',"済")'),
    ]
    for i, (label, f) in enumerate(summ, 4):
        dash.cell(i, 4, label)
        c = dash.cell(i, 5, f); c.number_format = '¥#,##0'
    dash.column_dimensions["A"].width = 16; dash.column_dimensions["D"].width = 24; dash.column_dimensions["E"].width = 16
    dash["A11"] = "※金額は税込。ステータス・入金はプルダウンで入力。行は400行目まで数式対応。"
    dash["A11"].font = Font(size=9, color="5A6E8C")

    out = os.path.join(HERE, "案件管理.xlsx")
    wb.save(out)
    print("✅ 生成:", out)


if __name__ == "__main__":
    build()
