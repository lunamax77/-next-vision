#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
御見積書 / 御請求書 ジェネレーター

既存テンプレート(templates/quote_template.xlsx, invoice_template.xlsx)を土台に、
ジョブJSONで指定した 顧客・期間・遊具 から xlsx を自動生成する。
テンプレの体裁・印影・数式(小計/消費税/合計)はそのまま保持される。

使い方:
    python generate.py jobs/sample_quote.json
    python generate.py jobs/sample_invoice.json

ジョブJSONの形式は jobs/sample_*.json と README.md を参照。
遊具名だけ指定すれば pricebook.json から単価を引いて金額を自動計算する
(初日満額・2日目以降50%)。unit_price / amount を明示すればそちらを優先。
"""
import json, sys, os, datetime
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))

# テンプレ種別ごとのセル位置マップ
LAYOUT = {
    "quote": {
        "template": "templates/quote_template.xlsx",
        "sheet": "御見積書11月",
        "date": "H1", "person": "H2",
        "customer": {"cell": "A6", "fmt": "{name} 御中"},
        "subject":  {"cell": "A8", "fmt": "件名：{subject}"},
        "period":   {"cell": "A15", "fmt": "実施期間：{period}"},
        "due": None,
        "rental_rows": [18, 19, 20, 21],       # 小計に含まれる行
        "transport_rows": [23, 24, 25],        # 小計 =SUM(G18:G21,G23:G25)
    },
    "invoice": {
        "template": "templates/invoice_template.xlsx",
        "sheet": "御請求書",
        "date": "H1", "person": "H2",
        "customer": {"cell": "A7", "fmt": "{name}　御中"},
        "subject":  {"cell": "A9", "fmt": "件名：{subject}"},
        "period":   {"cell": "A18", "fmt": "実施期間：{period}"},
        "due":      {"cell": "A15", "fmt": "お手数ですが、{due}までに下記の口座までお振込みをお願い致します。"},
        "rental_rows": [20, 21, 22, 23],
        "transport_rows": [25, 26, 27, 28, 29],  # 小計 =SUM(G20:G29)
    },
}

# 行内カラム
COL = {"name": "B", "qty": "D", "days": "E", "unit": "F", "amount": "G", "note": "H"}


def load_pricebook():
    with open(os.path.join(HERE, "pricebook.json"), encoding="utf-8") as f:
        return json.load(f)


def calc_amount(pb, name, qty, days):
    """pricebook から金額を自動計算。個別見積品は None を返す。"""
    it = pb["items"].get(name)
    if not it or it.get("quote_only") or not it.get("daily"):
        return None, None
    daily = it["daily"]
    r = pb["multiday_rule"]
    per_unit = daily * (r["first_day"] + r["extra_day"] * max(0, days - 1))
    return daily, round(per_unit * qty)


def resolve_items(pb, items):
    """ジョブの item 指定を {name,qty,days,unit,amount,note} に正規化。"""
    out = []
    for raw in items:
        name = raw.get("name") or raw.get("item")
        qty = raw.get("qty", 1)
        days = raw.get("days", 1)
        unit = raw.get("unit_price")
        amount = raw.get("amount")
        note = raw.get("note", "")
        if amount is None or unit is None:
            d, a = calc_amount(pb, name, qty, days)
            if unit is None:
                unit = d
            if amount is None:
                amount = a
        out.append({"name": name, "qty": qty, "days": days,
                    "unit": unit, "amount": amount, "note": note})
    return out


def write_row(ws, row, item):
    ws[f'{COL["name"]}{row}'] = item["name"]
    ws[f'{COL["qty"]}{row}'] = item["qty"]
    ws[f'{COL["days"]}{row}'] = item["days"]
    ws[f'{COL["unit"]}{row}'] = item["unit"] if item["unit"] is not None else "個別見積"
    if item["amount"] is not None:
        ws[f'{COL["amount"]}{row}'] = item["amount"]
    if item["note"]:
        ws[f'{COL["note"]}{row}'] = item["note"]


def generate(job):
    typ = job["type"]
    lay = LAYOUT[typ]
    pb = load_pricebook()

    wb = openpyxl.load_workbook(os.path.join(HERE, lay["template"]))
    ws = wb[lay["sheet"]]

    if job.get("issue_date"):
        d = job["issue_date"]
        try:  # "YYYY-MM-DD" は日付型に変換(テンプレの和暦表示書式が効く)
            y, m, dd = (int(x) for x in d.split("-"))
            ws[lay["date"]] = datetime.date(y, m, dd)
        except (ValueError, AttributeError):
            ws[lay["date"]] = d  # 未指定ならテンプレの=TODAY()のまま
    if job.get("person"):
        ws[lay["person"]] = job["person"]
    ws[lay["customer"]["cell"]] = lay["customer"]["fmt"].format(name=job["customer"])
    ws[lay["subject"]["cell"]] = lay["subject"]["fmt"].format(subject=job.get("subject", ""))
    ws[lay["period"]["cell"]] = lay["period"]["fmt"].format(period=job.get("period", ""))
    if lay["due"] and job.get("due_date"):
        ws[lay["due"]["cell"]] = lay["due"]["fmt"].format(due=job["due_date"])

    warnings = []
    rental = resolve_items(pb, job.get("rental_items", []))
    transport = resolve_items(pb, job.get("transport_items", []))

    for i, item in enumerate(rental):
        if i < len(lay["rental_rows"]):
            write_row(ws, lay["rental_rows"][i], item)
        else:
            warnings.append(f'レンタル明細が上限({len(lay["rental_rows"])}行)を超過: 「{item["name"]}」は未記載')
    for i, item in enumerate(transport):
        if i < len(lay["transport_rows"]):
            write_row(ws, lay["transport_rows"][i], item)
        else:
            warnings.append(f'運送費明細が上限({len(lay["transport_rows"])}行)を超過: 「{item["name"]}」は未記載')

    out = job["output"]
    if not os.path.isabs(out):
        out = os.path.join(HERE, out)
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    return out, warnings, rental, transport


def main():
    if len(sys.argv) < 2:
        print("usage: python generate.py <job.json>"); sys.exit(1)
    with open(sys.argv[1], encoding="utf-8") as f:
        job = json.load(f)
    out, warnings, rental, transport = generate(job)
    subtotal = sum(x["amount"] or 0 for x in rental + transport)
    print(f"✅ 生成: {out}")
    print(f"   種別: {'御見積書' if job['type']=='quote' else '御請求書'} / 宛先: {job['customer']}")
    print(f"   明細: レンタル{len(rental)}件 + 運送{len(transport)}件 / 小計(税別) ￥{subtotal:,} → 税込 ￥{round(subtotal*1.1):,}")
    for w in warnings:
        print(f"   ⚠ {w}")


if __name__ == "__main__":
    main()
